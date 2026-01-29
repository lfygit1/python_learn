# 下载三方库 openpyxl
# pip install openpyxl==3.1.2 -i https://pypi.tuna.tsinghua.edu.cn/simple/   没网也可下载
# 查看是否下载成功   pip show openpyxl

import openpyxl
def read_excl(a,b):  # a是路径加文件名，b是sheet页
    # 打开工作薄
    wb=openpyxl.load_workbook(a)
    # 选择sheet页
    st=wb[b]  # st就是当前的sheet页
    # 获取最大行
    mr=st.max_row
    print(mr)
    # 获取最大列
    mc=st.max_column
    print(mc)
    # 获取第一行的内容
    for i in range(1,mc+1):
        print(st.cell(1,i).value,end=' ')
    print()

    # 获取第一列的内容
    for i in range (1,mr+1):
        print(st.cell(i,1).value,end=' ')
    print()
    # 获取单元格的内容
    # for j in range(1,mr+1):
    #     for i in range (1,mc+1):
    #         print(st.cell(j, i).value,end=' ')
    #     print()

    # 把excl文件中的内容放到一个列表中，效果如下
    # [[],[],[],[]]
    da =[]
    for j in range(1,mr+1):
        xiao=[]
        for i in range(1, mc + 1):
            xiao.append(st.cell(j, i).value)
        da.append(xiao)
    return da

read_excl(r'D:\Python_learn\files\sbggg.xlsx','sheet1')
