# 题目 1：简单求和函数
# 需求：定义一个函数 get_sum，接收两个数字参数 a 和 b，返回这两个数的和；调用函数，传入参数 10 和 25，打印结果。
# 思路：基础函数定义，def 关键字 + 参数 +return 返回值，直接调用即可。
def get_sum(a,b):
    return a+b
# print(get_sum(10, 25))

# 题目 2：带默认参数的问候函数
# 需求：定义一个函数 greet，接收姓名（必选参数）和问候语（默认参数，默认值为你好），输出「问候语 + 姓名」；
# 调用 1：只传姓名小明，使用默认问候语；
# 调用 2：传姓名小红和问候语早上好，使用自定义问候语。
# 思路：默认参数放在必选参数后面，调用时可省略默认参数。
def greet (name,talk='hello'):
    return talk +name
# print(greet('小明'))
# print(greet('小红','早上好'))


# 题目 3：求列表平均值的函数
# 需求：定义一个函数 get_average，接收一个数字列表作为参数，计算并返回列表的平均值；
# 测试：传入列表 [80, 90, 75, 85]，打印平均值。
# 思路：用sum()求列表和，len()求列表长度，和 ÷ 长度即为平均值，注意返回数值。
li=[80, 90, 75, 85]
def get_average(*a):
        return sum(a)/len(a)
# print(get_average(*li))


# 题目 4：判断奇偶函数
# 需求：定义一个函数 is_even，接收一个整数参数 n，如果是偶数返回True，奇数返回False；
# 测试：分别传入 17 和 24，打印结果。
# 思路：用取模运算%，n % 2 == 0 即为偶数。

def is_even(n):
    if n %2==0:
        return True
    else:
        return False
# print(is_even(24))

# 题目 5：无返回值的打印函数
# 需求：定义一个函数 print_info，接收任意多个字符串参数（可变参数*args），依次打印每个参数，每个参数占一行；
# 测试：传入"Python", "函数", "基础练习"，查看打印结果。
def print_info(*args):
    for i in args:
        print(i)
# print_info("Python", "函数", "基础练习")


# 题目 6：基础类定义与实例化
# 需求：定义一个 Dog 类，包含实例属性：name（名字）、age（年龄）；包含实例方法：bark()，输出「名字 + 汪汪叫」；
# 1、创建Dog类的实例：名字旺财，年龄 3；
# 2、打印实例的名字和年龄；
# 3、调用实例的bark()方法。
# 思路：__init__方法初始化实例属性，方法中通过self访问属性。
class Dog():
    def __init__(self,name,age):
        self.name=name
        self.age=age
    def bark(self):
        return (f'一号狗子{self.name}，汪汪叫')
    def dog(self):
        return (f'一号狗子{self.name},今年{self.age}岁')

# dog1=Dog('旺财',3)
# print(dog1.bark())
# print(dog1.dog())





# 题目 7：简单购物类
# 需求：定义一个 Shopping 类，包含实例属性：total（总金额，初始值 0）；
# 包含两个方法：
# add_goods：接收一个数字参数（商品价格），将价格加到total上；
# show_total：无参数，打印「当前购物车总金额：XXX 元」；
# 测试：
# 1、创建Shopping实例；
# 2、依次添加商品：59 元、129 元、89 元；
# 3、调用show_total查看总金额。
class Shopping():
    def __init__(self,total=0):
        self.total=total
    def add_goods(self,price):
        self.total= int(price) +self.total
    def show_total(self):
        return (f'当前购物车总金额：{self.total}')
# goods1=Shopping()
# print(goods1.add_goods(59))
# print(goods1.add_goods(128))
# print(goods1.show_total())




# 题目 8：函数调用类的方法
# 需求：定义一个 Book 类，包含属性name（书名）、price（价格），方法get_info返回「书名：XXX，价格：XXX 元」；
# 定义一个函数 print_book_info，接收一个Book类的实例作为参数，调用实例的get_info方法并打印结果；
# 测试：创建Book实例（书名《Python 基础》，价格 39.9），调用print_book_info函数。
class Book():
    def __init__(self,name,price):
        self.name=name
        self.price=price
    def get_info(self):
        print(self.name,self.price)
    def print_book_info(self):
        return (self.get_info())

book1=Book('《Python 基础》',39.9)
# book1.print_book_info()





#
# 题目 9：JSON 文件写入与读取
# 需求：
# 1、定义一个字典 student，包含：name（张三）、age（20）、score（85）、hobbies（["篮球", "编程"]）；
# 2、将该字典写入 student.json 文件（格式化输出，缩进 2）；
# 3、读取 student.json 文件，解析为字典并打印；
# 4、提取并打印字典中的 hobbies 字段。
import json
student={
    'name':'张三',
    'age':20,
    'score':85,
    'hobbies':['篮球','编程']
    }

with open(r'D:\Python_learn\files\student.json','w',encoding='utf-8') as fp:
    json.dump(student,fp,ensure_ascii=False)

with open(r'D:\Python_learn\files\student.json','r',encoding='utf-8') as fp:
    s=json.load(fp)
    # print(s)
    # print(s["hobbies"])

# 题目 10：Excel → YAML 配置导出
# 需求：
# 1、读取 score.xlsx 文件的「成绩表」数据；
import openpyxl
duqu=openpyxl.load_workbook(r'D:\Python_learn\files\score.xlsx')  # 读取文件
sheet=duqu['Sheet1']  # 读取sheet页
hang=sheet.max_row
# print(hang)    # 获取最大行
lie=sheet.max_column
# print(lie)    # 获取最大列
wai =[]
for i in range(1,hang+1):
    nei=[]
    for j in range(1,lie+1):
        nei.append(sheet.cell(i,j).value)
    wai.append(nei)
print(wai)


# 提取所有学生的姓名和总分，构建字典（键：姓名，值：总分）；
dict1={}
for i in wai[1:]:
    name=i[0]
    score= i[1:]
    total=sum(score)
    dict1[name]=total
print(dict1)


