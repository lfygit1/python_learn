import openpyxl
def test_sum1():
    print('dfdfdsfds')

def test_read_excl():  # a是路径加文件名，b是sheet页
    # 打开工作薄
    a=r'D:\Python_learn\files\sbggg.xlsx'
    b='sheet1'
    wb=openpyxl.load_workbook(a)
    # 选择sheet页
    st=wb[b]  # st就是当前的sheet页
    # 获取最大行
    mr=st.max_row
    print(mr)
    # 获取最大列
    mc=st.max_column
    print(mc)
    # 获取第一行的内容
    for i in range(1,mc+1):
        print(st.cell(1,i).value,end=' ')
    print()

    # 获取第一列的内容
    for i in range (1,mr+1):
        print(st.cell(i,1).value,end=' ')
    print()
# 写文件
def test_wright_txt():  # file_path 指的是文件路径   b 指的是读或者写 c 指的是写入内容
    file_path=r'D:\Python_learn\files\wright.txt'
    b='a'
    c="['hello\n','hi\n','goodbye\n','hahahaha']"
    with open(file_path,b,encoding='utf-8') as f:
        f.writelines(c)
